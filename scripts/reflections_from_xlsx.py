#!/usr/bin/env python3
"""
Convert a Reflection Asset Library tracker (.xlsx) into typed TypeScript records.

    pip install openpyxl
    python3 scripts/reflections_from_xlsx.py Breaking_Chains_Tracker.xlsx > data/reflections.ts

Reads the columns named in the Brand Guide's Asset Library Standard. Nothing is invented:
every field written out is a column in the sheet. Rows without a Reflection ID are skipped.

To convert a second collection (Rooted, Waiting for the Bridegroom, ...), run this against its
tracker and pass --collection / --book-slug / --author, then paste the records into the array.
"""

import argparse
import json
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

# Column heading -> record field. The left side must match the sheet exactly.
COLUMNS = {
    "Reflection ID": "code",
    "Title": "title",
    "Reflection": "text",
    "Book": "book",
    "Chapter": "chapter",
    "Theme": "theme",
    "Emotion": "emotion",
    "Suggested Background": "background",
    "Suggested Colour Palette": "palette",
    "Tags": "tags",
    "Priority": "priority",
}

HEADER = """// AUTO-GENERATED — do not edit by hand.
// Source: {source}
// Regenerate: python3 scripts/reflections_from_xlsx.py <tracker.xlsx> > data/reflections.ts
//
// `artwork` is null until a real image exists. Until then each reflection renders as a plate
// tinted from its own Suggested Colour Palette (see lib/palettes.ts).

import type {{ Reflection }} from "@/lib/types";

export const reflections: Reflection[] = [
"""

FIELD_ORDER = [
    "id", "code", "title", "text", "author", "authorSlug", "collection",
    "book", "bookSlug", "chapter", "theme", "emotion", "background", "palette",
]


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("xlsx")
    p.add_argument("--sheet", default="Reflection Library")
    p.add_argument("--collection", default="breaking-chains")
    p.add_argument("--book-slug", default="breaking-chains")
    p.add_argument("--author", default="Nathlee R. Grant")
    p.add_argument("--author-slug", default="nathlee-r-grant")
    args = p.parse_args()

    ws = load_workbook(args.xlsx, read_only=True)[args.sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit(f"{args.xlsx}: '{args.sheet}' is empty")

    headings = [str(h).strip() if h else "" for h in rows[0]]
    missing = [h for h in COLUMNS if h not in headings]
    if missing:
        sys.exit(f"Missing required column(s): {', '.join(missing)}")
    at = {h: headings.index(h) for h in COLUMNS}

    out = []
    for row in rows[1:]:
        raw_code = row[at["Reflection ID"]]
        if not raw_code:
            continue
        code = str(raw_code).strip()

        def cell(name: str) -> str:
            v = row[at[name]]
            return str(v).strip() if v is not None else ""

        out.append({
            "id": code.lower(),
            "code": code,
            "title": cell("Title"),
            "text": cell("Reflection"),
            "author": args.author,
            "authorSlug": args.author_slug,
            "collection": args.collection,
            "book": cell("Book"),
            "bookSlug": args.book_slug,
            "chapter": cell("Chapter"),
            "theme": cell("Theme"),
            "emotion": cell("Emotion"),
            "background": cell("Suggested Background"),
            "palette": cell("Suggested Colour Palette"),
            "tags": [t.strip() for t in cell("Tags").split(",") if t.strip()],
            # Priority is stored as stars (★★★★☆). Count them.
            "priority": cell("Priority").count("\u2605"),
            "artworkFile": f"{code}.png",
        })

    if not out:
        sys.exit("No rows with a Reflection ID were found.")

    j = lambda v: json.dumps(v, ensure_ascii=False)
    buf = [HEADER.format(source=args.xlsx)]
    for r in out:
        buf.append("  {\n")
        for field in FIELD_ORDER:
            buf.append(f"    {field}: {j(r[field])},\n")
        buf.append(f"    tags: {j(r['tags'])},\n")
        buf.append(f"    priority: {r['priority']},\n")
        buf.append("    artwork: null,\n")
        buf.append(f"    artworkFile: {j(r['artworkFile'])},\n")
        buf.append("    publishedDate: null,\n")
        buf.append("  },\n")
    buf.append("];\n\nexport default reflections;\n")

    sys.stdout.write("".join(buf))
    print(f"{len(out)} reflections converted.", file=sys.stderr)


if __name__ == "__main__":
    main()
